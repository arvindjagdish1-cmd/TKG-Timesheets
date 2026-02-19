"""
Export services for generating XLSX and PDF files.

This module contains the business logic for:
- Generating individual timesheet XLSX files from templates
- Generating individual expense report XLSX files
- Converting XLSX to PDF using LibreOffice
- Merging PDFs into combined packs
- Creating ZIP bundles
"""

import os
import subprocess
import tempfile
from pathlib import Path
from datetime import timedelta
from decimal import Decimal

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def get_export_root():
    """Get the export root directory, creating if needed."""
    root = Path(settings.EXPORT_ROOT)
    root.mkdir(parents=True, exist_ok=True)
    return root


def generate_upload_xlsx(upload):
    """
    Generate a formatted XLSX from a TimesheetUpload's parsed_json data.

    Args:
        upload: TimesheetUpload model instance

    Returns:
        Path to generated file
    """
    from datetime import date as date_type
    from calendar import monthrange

    employee = upload.user
    year, month = upload.year, upload.month
    parsed = upload.parsed_json or {}
    time_data = parsed.get("time", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws["A1"] = "TKG Time Sheet"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Employee:"
    ws["B3"] = employee.get_full_name()
    ws["A4"] = "Period:"
    ws["B4"] = f"{year}-{month:02d}"
    ws["A5"] = "Status:"
    ws["B5"] = upload.get_status_display()

    def _write_half(start_row, half_key, half_label):
        half = time_data.get(half_key, {})
        lines = half.get("lines", [])
        daily_totals = half.get("daily_totals", {})

        if half_key == "first_half":
            dates = [date_type(year, month, d) for d in range(1, 16)]
        else:
            last_day = monthrange(year, month)[1]
            dates = [date_type(year, month, d) for d in range(16, last_day + 1)]

        row = start_row
        ws.cell(row=row, column=1, value=half_label).font = Font(bold=True, size=12)
        row += 1

        ws.cell(row=row, column=1, value="Charge Code").font = header_font_white
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Description").font = header_font_white
        ws.cell(row=row, column=2).fill = header_fill

        for i, d in enumerate(dates):
            cell = ws.cell(row=row, column=3 + i, value=d.strftime("%m/%d\n%a"))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(3 + i)].width = 8

        total_col = 3 + len(dates)
        ws.cell(row=row, column=total_col, value="Total").font = header_font_white
        ws.cell(row=row, column=total_col).fill = header_fill
        row += 1

        for line in lines:
            code = line.get("charge_code", "")
            label = line.get("label", code)
            ws.cell(row=row, column=1, value=code).border = border
            ws.cell(row=row, column=2, value=label).border = border

            daily = line.get("daily", {})
            line_total = Decimal("0")
            for i, d in enumerate(dates):
                hours = Decimal(str(daily.get(d.isoformat(), 0)))
                cell = ws.cell(
                    row=row, column=3 + i,
                    value=float(hours) if hours else None,
                )
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                line_total += hours

            ws.cell(row=row, column=total_col, value=float(line_total)).border = border
            ws.cell(row=row, column=total_col).font = Font(bold=True)
            row += 1

        ws.cell(row=row, column=1, value="Daily Total").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="f0f0f0", fill_type="solid"
        )
        grand = Decimal("0")
        for i, d in enumerate(dates):
            day_total = Decimal(str(daily_totals.get(d.isoformat(), 0)))
            cell = ws.cell(row=row, column=3 + i, value=float(day_total))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="f0f0f0", fill_type="solid")
            cell.border = border
            grand += day_total

        ws.cell(row=row, column=total_col, value=float(grand)).font = Font(
            bold=True, size=12
        )
        ws.cell(row=row, column=total_col).fill = PatternFill(
            start_color="c6f6d5", fill_type="solid"
        )
        row += 1
        return row

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28

    next_row = _write_half(8, "first_half", "Period One (1st – 15th)")
    next_row += 1
    _write_half(next_row, "second_half", "Period Two (16th – End)")

    row = ws.max_row + 2
    ws.cell(
        row=row, column=1,
        value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
    )

    output_dir = get_export_root() / f"{year}" / f"{month:02d}"
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_name = employee.get_full_name().replace(" ", "_").replace("/", "-")
    filename = f"timesheet_{year}_{month:02d}_{safe_name}.xlsx"
    output_path = output_dir / filename

    wb.save(output_path)
    return output_path


def generate_timesheet_xlsx(timesheet):
    """
    Generate XLSX for a single timesheet.

    Args:
        timesheet: Timesheet model instance

    Returns:
        Path to generated file
    """
    employee = timesheet.employee
    period = timesheet.period

    # Check if template exists
    template_path = Path(settings.TIMESHEET_TEMPLATE_PATH)
    if template_path.exists():
        wb = load_workbook(template_path)
        ws = wb.active
    else:
        # Create from scratch if no template
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"

    # Styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header section
    ws['A1'] = "TKG Time Sheet"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A3'] = "Employee:"
    ws['B3'] = employee.get_full_name()
    ws['A4'] = "Period:"
    ws['B4'] = period.display_name
    ws['A5'] = "Status:"
    ws['B5'] = timesheet.get_status_display()

    # Date range
    dates = []
    current = period.start_date
    while current <= period.end_date:
        dates.append(current)
        current += timedelta(days=1)

    # Column headers (dates)
    start_row = 8
    ws.cell(row=start_row, column=1, value="Charge Code").font = header_font_white
    ws.cell(row=start_row, column=1).fill = header_fill
    ws.cell(row=start_row, column=2, value="Description").font = header_font_white
    ws.cell(row=start_row, column=2).fill = header_fill

    for i, d in enumerate(dates):
        cell = ws.cell(row=start_row, column=3 + i, value=d.strftime("%m/%d\n%a"))
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(3 + i)].width = 8

    total_col = 3 + len(dates)
    ws.cell(row=start_row, column=total_col, value="Total").font = header_font_white
    ws.cell(row=start_row, column=total_col).fill = header_fill

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25

    # Data rows
    lines = timesheet.lines.select_related("charge_code").prefetch_related("entries").order_by("order", "id")
    row = start_row + 1

    for line in lines:
        ws.cell(row=row, column=1, value=line.charge_code.code).border = border
        label = line.label if line.label else line.charge_code.description
        ws.cell(row=row, column=2, value=label).border = border

        entry_map = {e.date: e.hours for e in line.entries.all()}
        line_total = Decimal("0")

        for i, d in enumerate(dates):
            hours = entry_map.get(d, Decimal("0"))
            cell = ws.cell(row=row, column=3 + i, value=float(hours) if hours else None)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            line_total += hours

        # Line total
        ws.cell(row=row, column=total_col, value=float(line_total)).border = border
        ws.cell(row=row, column=total_col).font = Font(bold=True)

        row += 1

    # Daily totals row
    ws.cell(row=row, column=1, value="Daily Total").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill(start_color="f0f0f0", fill_type="solid")

    grand_total = Decimal("0")
    for i, d in enumerate(dates):
        day_total = sum(
            e.hours for line in lines for e in line.entries.all() if e.date == d
        )
        cell = ws.cell(row=row, column=3 + i, value=float(day_total))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="f0f0f0", fill_type="solid")
        cell.border = border
        grand_total += day_total

    # Grand total
    ws.cell(row=row, column=total_col, value=float(grand_total)).font = Font(bold=True, size=12)
    ws.cell(row=row, column=total_col).fill = PatternFill(start_color="c6f6d5", fill_type="solid")

    # Footer
    row += 2
    ws.cell(row=row, column=1, value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    if timesheet.employee_notes:
        row += 1
        ws.cell(row=row, column=1, value=f"Notes: {timesheet.employee_notes}")

    # Save
    output_dir = get_export_root() / f"{period.year}" / f"{period.month:02d}"
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_name = employee.get_full_name().replace(" ", "_").replace("/", "-")
    filename = f"timesheet_{period.year}_{period.month:02d}_{period.half}_{safe_name}.xlsx"
    output_path = output_dir / filename

    wb.save(output_path)
    return output_path


def generate_expense_xlsx(expense_report):
    """
    Generate XLSX for a single expense report.

    Args:
        expense_report: ExpenseReport model instance

    Returns:
        Path to generated file
    """
    employee = expense_report.employee
    month = expense_report.month

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="38a169", end_color="38a169", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '"$"#,##0.00'

    # Header
    ws['A1'] = "TKG Expense Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A3'] = "Employee:"
    ws['B3'] = employee.get_full_name()
    ws['A4'] = "Month:"
    ws['B4'] = month.display_name
    ws['A5'] = "Status:"
    ws['B5'] = expense_report.get_status_display()

    # Expense Items Section
    start_row = 8
    headers = ["Date", "Category", "Description", "Client/Vendor", "Receipt", "Amount"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=start_row, column=i+1, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    items = expense_report.items.select_related("category").order_by("date")
    row = start_row + 1
    expense_total = Decimal("0")

    for item in items:
        ws.cell(row=row, column=1, value=item.date.strftime("%m/%d/%Y")).border = border
        ws.cell(row=row, column=2, value=item.category.name).border = border
        ws.cell(row=row, column=3, value=item.description).border = border
        ws.cell(row=row, column=4, value=item.client or "").border = border

        receipt_status = "Yes" if item.receipts.exists() else ("Paper" if item.paper_receipt_delivered else "No")
        ws.cell(row=row, column=5, value=receipt_status).border = border

        amount_cell = ws.cell(row=row, column=6, value=float(item.amount))
        amount_cell.border = border
        amount_cell.number_format = money_format
        amount_cell.alignment = Alignment(horizontal='right')

        expense_total += item.amount
        row += 1

    # Expense subtotal
    ws.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
    subtotal_cell = ws.cell(row=row, column=6, value=float(expense_total))
    subtotal_cell.font = Font(bold=True)
    subtotal_cell.number_format = money_format
    subtotal_cell.border = border

    # Mileage Section
    row += 2
    ws.cell(row=row, column=1, value="Mileage Reimbursement").font = Font(bold=True, size=12)
    row += 1

    mileage_headers = ["Date", "Description", "Miles", "Rate", "Amount"]
    for i, h in enumerate(mileage_headers):
        cell = ws.cell(row=row, column=i+1, value=h)
        cell.font = header_font_white
        cell.fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
        cell.border = border

    entries = expense_report.mileage_entries.order_by("date")
    row += 1
    mileage_total = Decimal("0")

    for entry in entries:
        ws.cell(row=row, column=1, value=entry.date.strftime("%m/%d/%Y")).border = border
        ws.cell(row=row, column=2, value=entry.description).border = border
        ws.cell(row=row, column=3, value=float(entry.miles)).border = border

        rate_cell = ws.cell(row=row, column=4, value=float(entry.rate))
        rate_cell.border = border
        rate_cell.number_format = '"$"0.000'

        amount_cell = ws.cell(row=row, column=5, value=float(entry.total_amount))
        amount_cell.border = border
        amount_cell.number_format = money_format
        amount_cell.alignment = Alignment(horizontal='right')

        mileage_total += entry.total_amount
        row += 1

    # Mileage subtotal
    ws.cell(row=row, column=4, value="Subtotal:").font = Font(bold=True)
    mileage_subtotal = ws.cell(row=row, column=5, value=float(mileage_total))
    mileage_subtotal.font = Font(bold=True)
    mileage_subtotal.number_format = money_format
    mileage_subtotal.border = border

    # Grand Total
    row += 2
    ws.cell(row=row, column=4, value="GRAND TOTAL:").font = Font(bold=True, size=14)
    grand_total_cell = ws.cell(row=row, column=5, value=float(expense_total + mileage_total))
    grand_total_cell.font = Font(bold=True, size=14)
    grand_total_cell.number_format = money_format
    grand_total_cell.fill = PatternFill(start_color="c6f6d5", fill_type="solid")
    grand_total_cell.border = border

    # Footer
    row += 2
    ws.cell(row=row, column=1, value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    if expense_report.employee_notes:
        row += 1
        ws.cell(row=row, column=1, value=f"Notes: {expense_report.employee_notes}")

    # Save
    output_dir = get_export_root() / f"{month.year}" / f"{month.month:02d}"
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_name = employee.get_full_name().replace(" ", "_").replace("/", "-")
    filename = f"expenses_{month.year}_{month.month:02d}_{safe_name}.xlsx"
    output_path = output_dir / filename

    wb.save(output_path)
    return output_path


def convert_xlsx_to_pdf(xlsx_path):
    """
    Convert XLSX to PDF using LibreOffice headless.

    Args:
        xlsx_path: Path to XLSX file

    Returns:
        Path to generated PDF
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    output_dir = xlsx_path.parent

    # LibreOffice headless conversion
    cmd = [
        "libreoffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(output_dir),
        str(xlsx_path),
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")

    # The output PDF has the same name but .pdf extension
    pdf_path = xlsx_path.with_suffix(".pdf")
    if not pdf_path.exists():
        raise FileNotFoundError(f"Expected PDF not found: {pdf_path}")

    return pdf_path


def merge_pdfs(pdf_paths, output_path):
    """
    Merge multiple PDFs into a single file.

    Args:
        pdf_paths: List of paths to PDF files
        output_path: Path for the merged output

    Returns:
        Path to merged PDF
    """
    from pypdf import PdfWriter

    writer = PdfWriter()

    for pdf_path in pdf_paths:
        if Path(pdf_path).exists():
            writer.append(str(pdf_path))

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "wb") as f:
        writer.write(f)

    return output_path


def create_zip_bundle(file_paths, output_path):
    """
    Create a ZIP archive containing all specified files.

    Args:
        file_paths: List of paths to include
        output_path: Path for the ZIP file

    Returns:
        Path to ZIP file
    """
    import zipfile

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in file_paths:
            file_path = Path(file_path)
            if file_path.exists():
                zf.write(file_path, file_path.name)

    return output_path
