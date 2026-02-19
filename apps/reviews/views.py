"""
Office Manager review dashboard and approval workflow.
Managing Partner and Payroll Partner views.
"""
import csv
from collections import defaultdict
from datetime import timedelta, date
from decimal import Decimal
from io import BytesIO, StringIO

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType
from django.conf import settings
from calendar import monthrange

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import json

from apps.timesheets.models import Timesheet, TimesheetLine, TimeEntry, ChargeCode, TimesheetUpload, ClientMapping
from apps.expenses.models import ExpenseReport, ExpenseCategory, ExpenseItem
from apps.periods.models import TimesheetPeriod, ExpenseMonth
from apps.accounts.models import User, EmployeeProfile
from .models import ReviewAction, ReviewComment, ManagingPartnerLayout


def _parse_month_param(value):
    if not value:
        today = timezone.now().date()
        return today.year, today.month
    try:
        year_str, month_str = value.split("-")
        return int(year_str), int(month_str)
    except ValueError:
        today = timezone.now().date()
        return today.year, today.month


def _month_label(year, month):
    return date(year, month, 1).strftime("%B %Y")


def _month_options(raw_qs, selected_year, selected_month):
    """Build pre-formatted month options for template dropdowns."""
    sel = f"{selected_year}-{selected_month:02d}"
    return [
        {
            "value": f"{y}-{m:02d}",
            "label": date(y, m, 1).strftime("%B %Y"),
            "selected": f"{y}-{m:02d}" == sel,
        }
        for y, m in raw_qs
    ]


def _active_employees():
    return User.objects.filter(
        is_active=True,
    ).select_related("profile").distinct().order_by("last_name", "first_name")


def _latest_upload_for_user(user, year, month):
    return (
        TimesheetUpload.objects.filter(user=user, year=year, month=month)
        .order_by("-uploaded_at")
        .first()
    )


def office_manager_required(view_func):
    """Decorator to require office_manager or higher role."""
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("account_login")
        if not request.user.groups.filter(
            name__in=["office_manager", "managing_partner", "payroll_partner"]
        ).exists() and not request.user.is_superuser:
            return HttpResponseForbidden("Access denied. Office Manager role required.")
        return view_func(request, *args, **kwargs)
    return wrapped


def managing_partner_required(view_func):
    """Decorator to require managing_partner, payroll_partner, or partners role."""
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("account_login")
        if not request.user.groups.filter(
            name__in=["managing_partner", "payroll_partner", "partners"]
        ).exists() and not request.user.is_superuser:
            return HttpResponseForbidden("Access denied. Partner role required.")
        return view_func(request, *args, **kwargs)
    return wrapped


def payroll_partner_required(view_func):
    """Decorator to require payroll_partner role."""
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("account_login")
        if not request.user.groups.filter(
            name__in=["payroll_partner", "managing_partner"]
        ).exists() and not request.user.is_superuser:
            return HttpResponseForbidden("Access denied. Payroll Partner role required.")
        return view_func(request, *args, **kwargs)
    return wrapped


def managing_partner_only_required(view_func):
    """Restrict access to managing_partner group or superuser only."""
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("account_login")
        if not request.user.groups.filter(
            name="managing_partner"
        ).exists() and not request.user.is_superuser:
            return HttpResponseForbidden("Access denied. Managing Partner role required.")
        return view_func(request, *args, **kwargs)
    return wrapped


@login_required
@office_manager_required
def review_dashboard(request):
    """Office Manager dashboard: submission status overview."""
    year, month = _parse_month_param(request.GET.get("month"))
    status_filter = request.GET.get("status", "ALL")
    status_options = ["ALL", "DRAFT", "SUBMITTED", "RETURNED", "APPROVED", "MISSING"]

    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:6]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    rows = []
    pending_count = 0
    active_employees = _active_employees()

    for emp in active_employees:
        upload = _latest_upload_for_user(emp, year, month)
        if not upload:
            status = "MISSING"
            hours = 0
            expenses = 0
            error_state = "missing"
        else:
            status = upload.status
            summary = upload.parsed_json or {}
            hours = (
                (summary.get("time", {}).get("first_half", {}).get("total_hours") or 0)
                + (summary.get("time", {}).get("second_half", {}).get("total_hours") or 0)
            )
            expenses = summary.get("expenses", {}).get("total_expenses", 0)
            if upload.has_blocking_errors:
                error_state = "blocking"
            elif any(issue.get("severity") == "WARN" for issue in upload.errors_json):
                error_state = "warning"
            else:
                error_state = "ok"
            if upload.status == TimesheetUpload.Status.SUBMITTED:
                pending_count += 1

        if status_filter != "ALL" and status != status_filter:
            continue

        rows.append({
            "employee": emp,
            "upload": upload,
            "status": status,
            "total_hours": hours,
            "total_expenses": expenses,
            "error_state": error_state,
        })

    context = {
        "rows": rows,
        "month_options": _month_options(raw_months, year, month),
        "selected_month_label": _month_label(year, month),
        "status_filter": status_filter,
        "status_options": status_options,
        "pending_count": pending_count,
    }
    return render(request, "reviews/office_dashboard.html", context)


@login_required
@office_manager_required
def pending_reviews(request):
    """List all pending submissions awaiting review."""
    pending_uploads = TimesheetUpload.objects.filter(
        status=TimesheetUpload.Status.SUBMITTED
    ).select_related("user").order_by("uploaded_at")

    context = {
        "pending_uploads": pending_uploads,
    }
    return render(request, "reviews/pending.html", context)


@login_required
@office_manager_required
def office_employee_detail(request, user_id, year, month):
    employee = get_object_or_404(User, pk=user_id)
    upload = (
        TimesheetUpload.objects.filter(user=employee, year=year, month=month)
        .order_by("-uploaded_at")
        .first()
    )
    if not upload:
        messages.error(request, "No upload found for this employee.")
        return redirect("reviews:dashboard")

    errors = [i for i in upload.errors_json if i.get("severity") == "ERROR"]
    warnings = [i for i in upload.errors_json if i.get("severity") == "WARN"]

    context = {
        "employee": employee,
        "upload": upload,
        "summary": upload.parsed_json or {},
        "errors": errors,
        "warnings": warnings,
    }
    return render(request, "reviews/employee_detail_readonly.html", context)


@login_required
@office_manager_required
@require_POST
def office_return_upload(request, pk):
    upload = get_object_or_404(TimesheetUpload, pk=pk)
    comment = request.POST.get("comment", "").strip()
    upload.status = TimesheetUpload.Status.RETURNED
    upload.reviewer_comment = comment
    upload.reviewed_by = request.user
    upload.reviewed_at = timezone.now()
    upload.save(update_fields=["status", "reviewer_comment", "reviewed_by", "reviewed_at", "updated_at"])
    messages.success(request, f"Upload returned for {upload.user.get_full_name()}.")
    return redirect("reviews:office_employee_detail", user_id=upload.user_id, year=upload.year, month=upload.month)


@login_required
@office_manager_required
@require_POST
def office_approve_upload(request, pk):
    upload = get_object_or_404(TimesheetUpload, pk=pk)
    comment = request.POST.get("comment", "").strip()
    upload.status = TimesheetUpload.Status.APPROVED
    upload.reviewer_comment = comment
    upload.reviewed_by = request.user
    upload.reviewed_at = timezone.now()
    upload.save(update_fields=["status", "reviewer_comment", "reviewed_by", "reviewed_at", "updated_at"])
    messages.success(request, f"Upload approved for {upload.user.get_full_name()}.")
    return redirect("reviews:office_employee_detail", user_id=upload.user_id, year=upload.year, month=upload.month)


@login_required
@office_manager_required
def review_timesheet(request, pk):
    """Review a submitted timesheet."""
    timesheet = get_object_or_404(
        Timesheet.objects.select_related("employee", "period", "employee__profile"),
        pk=pk,
    )

    lines = timesheet.lines.select_related("charge_code").prefetch_related("entries").all()

    # Get date range
    from apps.timesheets.views import _get_period_dates
    dates = _get_period_dates(timesheet.period)

    # Build entry data
    entry_data = {}
    for line in lines:
        entry_data[line.id] = {entry.date: entry.hours for entry in line.entries.all()}

    # Get review history
    ct = ContentType.objects.get_for_model(Timesheet)
    actions = ReviewAction.objects.filter(
        content_type=ct, object_id=timesheet.pk
    ).select_related("actor").order_by("-created_at")

    comments = ReviewComment.objects.filter(
        content_type=ct, object_id=timesheet.pk
    ).select_related("author").order_by("created_at")

    context = {
        "timesheet": timesheet,
        "lines": lines,
        "dates": dates,
        "entry_data": entry_data,
        "actions": actions,
        "comments": comments,
    }
    return render(request, "reviews/review_timesheet.html", context)


@login_required
@office_manager_required
@require_POST
def approve_timesheet(request, pk):
    """Approve a submitted timesheet."""
    timesheet = get_object_or_404(Timesheet, pk=pk)

    if timesheet.status != Timesheet.Status.SUBMITTED:
        messages.error(request, "Only submitted timesheets can be approved.")
        return redirect("reviews:review_timesheet", pk=pk)

    try:
        with transaction.atomic():
            timesheet.approve(request.user)
            ReviewAction.log_action(
                timesheet,
                ReviewAction.ActionType.APPROVED,
                request.user,
                request.POST.get("comment", "")
            )
        messages.success(request, f"Timesheet for {timesheet.employee.get_full_name()} approved.")
    except Exception as e:
        messages.error(request, str(e))

    # Redirect to next pending or dashboard
    next_pending = Timesheet.objects.filter(status=Timesheet.Status.SUBMITTED).first()
    if next_pending:
        return redirect("reviews:review_timesheet", pk=next_pending.pk)
    return redirect("reviews:dashboard")


@login_required
@office_manager_required
@require_POST
def return_timesheet(request, pk):
    """Return a timesheet for revision."""
    timesheet = get_object_or_404(Timesheet, pk=pk)

    if timesheet.status != Timesheet.Status.SUBMITTED:
        messages.error(request, "Only submitted timesheets can be returned.")
        return redirect("reviews:review_timesheet", pk=pk)

    comment = request.POST.get("comment", "").strip()
    if not comment:
        messages.error(request, "Please provide a reason for returning the timesheet.")
        return redirect("reviews:review_timesheet", pk=pk)

    try:
        with transaction.atomic():
            timesheet.return_for_revision(comment)
            ReviewAction.log_action(
                timesheet,
                ReviewAction.ActionType.RETURNED,
                request.user,
                comment
            )
        messages.success(request, f"Timesheet returned to {timesheet.employee.get_full_name()}.")
    except Exception as e:
        messages.error(request, str(e))

    return redirect("reviews:pending")


@login_required
@office_manager_required
def review_expense(request, pk):
    """Review a submitted expense report."""
    report = get_object_or_404(
        ExpenseReport.objects.select_related("employee", "month", "employee__profile"),
        pk=pk,
    )

    items = report.items.select_related("category").prefetch_related("receipts").order_by("date")
    mileage_entries = report.mileage_entries.order_by("date")

    # Get review history
    ct = ContentType.objects.get_for_model(ExpenseReport)
    actions = ReviewAction.objects.filter(
        content_type=ct, object_id=report.pk
    ).select_related("actor").order_by("-created_at")

    comments = ReviewComment.objects.filter(
        content_type=ct, object_id=report.pk
    ).select_related("author").order_by("created_at")

    context = {
        "report": report,
        "items": items,
        "mileage_entries": mileage_entries,
        "actions": actions,
        "comments": comments,
    }
    return render(request, "reviews/review_expense.html", context)


@login_required
@office_manager_required
@require_POST
def approve_expense(request, pk):
    """Approve a submitted expense report."""
    report = get_object_or_404(ExpenseReport, pk=pk)

    if report.status != ExpenseReport.Status.SUBMITTED:
        messages.error(request, "Only submitted expense reports can be approved.")
        return redirect("reviews:review_expense", pk=pk)

    try:
        with transaction.atomic():
            report.approve(request.user)
            ReviewAction.log_action(
                report,
                ReviewAction.ActionType.APPROVED,
                request.user,
                request.POST.get("comment", "")
            )
        messages.success(request, f"Expense report for {report.employee.get_full_name()} approved.")
    except Exception as e:
        messages.error(request, str(e))

    next_pending = ExpenseReport.objects.filter(status=ExpenseReport.Status.SUBMITTED).first()
    if next_pending:
        return redirect("reviews:review_expense", pk=next_pending.pk)
    return redirect("reviews:dashboard")


@login_required
@office_manager_required
@require_POST
def return_expense(request, pk):
    """Return an expense report for revision."""
    report = get_object_or_404(ExpenseReport, pk=pk)

    if report.status != ExpenseReport.Status.SUBMITTED:
        messages.error(request, "Only submitted expense reports can be returned.")
        return redirect("reviews:review_expense", pk=pk)

    comment = request.POST.get("comment", "").strip()
    if not comment:
        messages.error(request, "Please provide a reason for returning the expense report.")
        return redirect("reviews:review_expense", pk=pk)

    try:
        with transaction.atomic():
            report.return_for_revision(comment)
            ReviewAction.log_action(
                report,
                ReviewAction.ActionType.RETURNED,
                request.user,
                comment
            )
        messages.success(request, f"Expense report returned to {report.employee.get_full_name()}.")
    except Exception as e:
        messages.error(request, str(e))

    return redirect("reviews:pending")


@login_required
@office_manager_required
@require_POST
def add_comment(request, content_type, pk):
    """Add a comment to a timesheet or expense report."""
    if content_type == "timesheet":
        obj = get_object_or_404(Timesheet, pk=pk)
    elif content_type == "expense":
        obj = get_object_or_404(ExpenseReport, pk=pk)
    else:
        return HttpResponse("Invalid content type", status=400)

    text = request.POST.get("comment", "").strip()
    is_internal = request.POST.get("is_internal") == "on"

    if not text:
        return HttpResponse("Comment cannot be empty", status=400)

    ct = ContentType.objects.get_for_model(obj)
    ReviewComment.objects.create(
        content_type=ct,
        object_id=obj.pk,
        author=request.user,
        text=text,
        is_internal=is_internal,
    )

    if content_type == "timesheet":
        return redirect("reviews:review_timesheet", pk=pk)
    else:
        return redirect("reviews:review_expense", pk=pk)


# =============================================================================
# MANAGING PARTNER VIEWS
# =============================================================================

def _get_period_dates(period):
    """Get list of dates for a timesheet period."""
    dates = []
    current = period.start_date
    while current <= period.end_date:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def _get_daily_hours(timesheet, dates):
    """Get hours by date for a timesheet."""
    daily = {d: Decimal("0") for d in dates}
    for line in timesheet.lines.prefetch_related("entries").all():
        for entry in line.entries.all():
            if entry.date in daily:
                daily[entry.date] += entry.hours
    return daily


def _check_flags(daily_hours):
    """
    Check for flags:
    - Incomplete days (less than 8 hours on weekdays)
    - More than 10 hours per day for more than 2 days in a week
    """
    flags = {
        "incomplete_days": [],
        "excessive_hours_weeks": [],
    }
    
    # Group by ISO week
    weeks = defaultdict(list)
    
    for date, hours in daily_hours.items():
        # Skip weekends
        if date.weekday() >= 5:
            continue
        
        # Check incomplete (less than 8 hours on weekday)
        if hours < 8:
            flags["incomplete_days"].append({
                "date": date,
                "hours": hours,
            })
        
        # Track days with >10 hours by week
        week_key = date.isocalendar()[:2]  # (year, week)
        if hours > 10:
            weeks[week_key].append(date)
    
    # Check if any week has more than 2 days with >10 hours
    for week_key, dates in weeks.items():
        if len(dates) > 2:
            flags["excessive_hours_weeks"].append({
                "week": week_key,
                "dates": dates,
                "count": len(dates),
            })
    
    return flags


@login_required
@managing_partner_required
def managing_partner_dashboard(request):
    """Managing Partner dashboard with period selection."""
    year, month = _parse_month_param(request.GET.get("month"))
    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:12]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    context = {
        "month_options": _month_options(raw_months, year, month),
        "selected_month": f"{year}-{month:02d}",
        "selected_month_label": _month_label(year, month),
    }
    return render(request, "reviews/managing_partner/dashboard.html", context)


@login_required
@managing_partner_required
def daily_summary(request, period_id=None):
    """
    By-day timesheet summary for a month (two half tables).
    """
    year, month = _parse_month_param(request.GET.get("month"))
    last_day = monthrange(year, month)[1]
    first_dates = [date(year, month, d) for d in range(1, 16)]
    second_dates = [date(year, month, d) for d in range(16, last_day + 1)]

    employees = _active_employees()
    high_hours = Decimal(str(getattr(settings, "HIGH_HOURS_THRESHOLD", 10)))

    def build_rows(dates, half_key):
        rows = []
        for emp in employees:
            upload = _latest_upload_for_user(emp, year, month)
            if not upload:
                rows.append({
                    "employee": emp,
                    "daily_totals": {d: None for d in dates},
                    "total_hours": Decimal("0"),
                    "missing": True,
                    "flags": {"high_hours": set(), "weekly_hours_flag": set()},
                })
                continue

            daily_map = upload.parsed_json.get("time", {}).get(half_key, {}).get("daily_totals", {})
            daily_totals = {d: Decimal(str(daily_map.get(d.isoformat(), 0))) for d in dates}

            flags = {"high_hours": set(), "weekly_hours_flag": set()}

            all_daily = {}
            for half in ("first_half", "second_half"):
                half_map = upload.parsed_json.get("time", {}).get(half, {}).get("daily_totals", {})
                for key, value in half_map.items():
                    all_daily[key] = Decimal(str(value))

            weekly = defaultdict(list)
            for day_str, hours in all_daily.items():
                day = date.fromisoformat(day_str)
                if day.weekday() < 5 and hours >= high_hours:
                    weekly[day.isocalendar()[:2]].append(day)
            for week_days in weekly.values():
                if len(week_days) >= 3:
                    flags["high_hours"].update(week_days)

            weekly_totals_map = defaultdict(Decimal)
            weekdays_per_week = defaultdict(int)
            for day_str, hrs in all_daily.items():
                day = date.fromisoformat(day_str)
                if day.weekday() >= 5:
                    continue
                week_key = day.isocalendar()[:2]
                weekly_totals_map[week_key] += Decimal(str(hrs))
                weekdays_per_week[week_key] += 1
            for week_key, total_hrs in weekly_totals_map.items():
                n_days = weekdays_per_week.get(week_key, 0)
                if n_days < 3:
                    continue
                if total_hrs < 35 or total_hrs > 55:
                    for day_str in all_daily.keys():
                        day = date.fromisoformat(day_str)
                        if day.isocalendar()[:2] == week_key:
                            flags["weekly_hours_flag"].add(day)

            total = sum(daily_totals.values())
            rows.append({
                "employee": emp,
                "daily_totals": daily_totals,
                "total_hours": total,
                "missing": False,
                "flags": flags,
            })
        return rows

    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:12]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    from django.urls import reverse
    export_url = reverse("reviews:partner_export_xlsx", args=[year, month])

    context = {
        "month_label": _month_label(year, month),
        "month_options": _month_options(raw_months, year, month),
        "first_dates": first_dates,
        "second_dates": second_dates,
        "first_rows": build_rows(first_dates, "first_half"),
        "second_rows": build_rows(second_dates, "second_half"),
        "export_url": export_url,
    }
    return render(request, "reviews/managing_partner/daily_summary.html", context)


@login_required
@managing_partner_required
def category_summary(request, period_id=None):
    """
    Category summary for a month (two half tables).
    """
    year, month = _parse_month_param(request.GET.get("month"))
    employees = _active_employees()

    client_labels = {}
    client_codes = set()

    def collect_client_labels(upload, half_key):
        for line in upload.parsed_json.get("time", {}).get(half_key, {}).get("lines", []):
            if line.get("group") == "client" and line.get("charge_code"):
                code = line.get("charge_code")
                label = line.get("label") or code
                client_codes.add(code)
                if code not in client_labels:
                    client_labels[code] = label

    uploads_by_user = {}
    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        if upload:
            uploads_by_user[emp.id] = upload
            collect_client_labels(upload, "first_half")
            collect_client_labels(upload, "second_half")

    marketing_rows = [
        ("Marketing - General/Other", "GEN"),
        ("Chicago - Strategics", "CHI-STRAT"),
        ("Chicago - Banks", "CHI-BNK"),
        ("Chicago - Existing", "CHI-EXST"),
        ("Chicago - PE", "CHI-PEG"),
        ("Atlanta - Strategics", "ATL-STRAT"),
        ("Atlanta - Banks", "ATL-BNK"),
        ("Atlanta - Existing", "ATL-EXST"),
        ("Atlanta - PE", "ATL-PEG"),
        ("Los Angeles - Strategics", "LAX-STRAT"),
        ("Los Angeles - Banks", "LAX-BNK"),
        ("Los Angeles - Existing", "LAX-EXST"),
        ("Los Angeles - PE", "LAX-PEG"),
    ]

    other_rows = [
        ("Administration", "ADM"),
        ("Recruiting", "REC"),
        ("Training", "TRN"),
        ("Meetings", "MTG"),
        ("PTO", "PTO"),
        ("Other Paid Time Off", "HOL+OFF"),
    ]

    client_order = list(
        ClientMapping.objects.filter(active=True).order_by("sort_order", "display_name")
    )
    client_order_codes = [c.code for c in client_order]

    def build_matrix(half_key):
        matrix = []
        employee_totals = defaultdict(Decimal)

        ordered_codes = [code for code in client_order_codes if code in client_codes and code and code != "0"]
        ordered_codes += sorted([code for code in client_codes if code not in ordered_codes and code and code != "0"])

        for code in ordered_codes:
            mapping = next((c for c in client_order if c.code == code), None)
            label = mapping.display_name if mapping else client_labels.get(code, code)
            row = {"label": label, "code": code, "group": "client", "values": {}}
            total = Decimal("0")
            for emp in employees:
                upload = uploads_by_user.get(emp.id)
                hours = Decimal("0")
                if upload:
                    totals = upload.parsed_json.get("time", {}).get(half_key, {}).get("totals_by_client_code", {})
                    hours = Decimal(str(totals.get(code, 0)))
                row["values"][emp.id] = hours
                total += hours
                employee_totals[emp.id] += hours
            row["total"] = total
            matrix.append(row)

        matrix.append({"label": "Total Chargeable", "group": "total", "values": {}, "total": None})

        for label, code in marketing_rows:
            row = {"label": label, "code": code, "group": "marketing", "values": {}}
            total = Decimal("0")
            for emp in employees:
                upload = uploads_by_user.get(emp.id)
                hours = Decimal("0")
                if upload:
                    totals = upload.parsed_json.get("time", {}).get(half_key, {}).get("totals_by_marketing_bucket", {})
                    hours = Decimal(str(totals.get(code, 0)))
                row["values"][emp.id] = hours
                total += hours
                employee_totals[emp.id] += hours
            row["total"] = total
            matrix.append(row)

        matrix.append({"label": "Total Marketing", "group": "total", "values": {}, "total": None})

        for label, code in other_rows:
            row = {"label": label, "code": code, "group": "other", "values": {}}
            total = Decimal("0")
            for emp in employees:
                upload = uploads_by_user.get(emp.id)
                hours = Decimal("0")
                if upload:
                    totals = upload.parsed_json.get("time", {}).get(half_key, {}).get("totals_by_other_hours", {})
                    if code == "HOL+OFF":
                        hours = Decimal(str(totals.get("HOL", 0))) + Decimal(str(totals.get("OFF", 0)))
                    else:
                        hours = Decimal(str(totals.get(code, 0)))
                row["values"][emp.id] = hours
                total += hours
                employee_totals[emp.id] += hours
            row["total"] = total
            matrix.append(row)

        matrix.append({"label": "Total Other Hours", "group": "total", "values": {}, "total": None})
        matrix.append({"label": "Total Hours", "group": "grand_total", "values": {}, "total": None})

        return matrix, employee_totals

    first_matrix, first_totals = build_matrix("first_half")
    second_matrix, second_totals = build_matrix("second_half")

    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:12]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    from django.urls import reverse
    export_url = reverse("reviews:partner_export_xlsx", args=[year, month])

    context = {
        "employees": employees,
        "month_label": _month_label(year, month),
        "month_options": _month_options(raw_months, year, month),
        "first_matrix": first_matrix,
        "second_matrix": second_matrix,
        "first_totals": first_totals,
        "second_totals": second_totals,
        "export_url": export_url,
    }
    return render(request, "reviews/managing_partner/category_summary.html", context)


# =============================================================================
# EMPLOYEE SUMMARY VIEWS (Hours + Expenses)
# =============================================================================

_EMPLOYEE_MARKETING_CODES = [
    ("GEN", "Marketing - General/Other"),
    ("CHI-STRAT", "Chicago - Strategics"),
    ("CHI-BNK", "Chicago - Banks"),
    ("CHI-EXST", "Chicago - Existing"),
    ("CHI-PEG", "Chicago - PE"),
    ("ATL-STRAT", "Atlanta - Strategics"),
    ("ATL-BNK", "Atlanta - Banks"),
    ("ATL-EXST", "Atlanta - Existing"),
    ("ATL-PEG", "Atlanta - PE"),
    ("LAX-STRAT", "Los Angeles - Strategics"),
    ("LAX-BNK", "Los Angeles - Banks"),
    ("LAX-EXST", "Los Angeles - Existing"),
    ("LAX-PEG", "Los Angeles - PE"),
]

_EMPLOYEE_OTHER_CODES = [
    ("ADM", "Administration"),
    ("REC", "Recruiting"),
    ("TRN", "Training"),
    ("MTG", "Meetings"),
    ("PTO", "PTO"),
    ("HOL+OFF", "Other Paid Time Off"),
]


def _build_employee_hours_half(employees, uploads_by_user, half_key):
    """Build employee-hours matrix for one half-month."""
    columns = [("client", "Client Hours")]
    columns += [("mktg:" + code, label) for code, label in _EMPLOYEE_MARKETING_CODES]
    columns += [("other:" + code, label) for code, label in _EMPLOYEE_OTHER_CODES]

    rows = []
    column_totals = {col_id: Decimal("0") for col_id, _ in columns}

    for emp in employees:
        upload = uploads_by_user.get(emp.id)
        row_values = {}
        row_total = Decimal("0")

        if upload:
            time_data = upload.parsed_json.get("time", {}).get(half_key, {})

            client_totals = time_data.get("totals_by_client_code", {})
            row_values["client"] = sum(Decimal(str(v)) for v in client_totals.values())

            mktg_totals = time_data.get("totals_by_marketing_bucket", {})
            for code, _ in _EMPLOYEE_MARKETING_CODES:
                row_values["mktg:" + code] = Decimal(str(mktg_totals.get(code, 0)))

            other_totals = time_data.get("totals_by_other_hours", {})
            for code, _ in _EMPLOYEE_OTHER_CODES:
                if code == "HOL+OFF":
                    row_values["other:" + code] = (
                        Decimal(str(other_totals.get("HOL", 0)))
                        + Decimal(str(other_totals.get("OFF", 0)))
                    )
                else:
                    row_values["other:" + code] = Decimal(str(other_totals.get(code, 0)))

        for col_id, _ in columns:
            v = row_values.get(col_id, Decimal("0"))
            column_totals[col_id] += v
            row_total += v

        rows.append({
            "employee": emp,
            "values": row_values,
            "total": row_total,
            "missing": upload is None,
        })

    grand_total = sum(column_totals.values())
    return columns, rows, column_totals, grand_total


@login_required
@managing_partner_required
def employee_summary(request):
    """
    Employee hours summary split by half-month: employees as rows,
    aggregated Client Hours + individual non-client charge codes as columns.
    """
    year, month = _parse_month_param(request.GET.get("month"))
    employees = _active_employees()

    uploads_by_user = {}
    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        if upload:
            uploads_by_user[emp.id] = upload

    first_columns, first_rows, first_totals, first_grand = _build_employee_hours_half(
        employees, uploads_by_user, "first_half"
    )
    second_columns, second_rows, second_totals, second_grand = _build_employee_hours_half(
        employees, uploads_by_user, "second_half"
    )

    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:12]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    from django.urls import reverse
    export_url = reverse("reviews:partner_export_xlsx", args=[year, month])

    context = {
        "month_label": _month_label(year, month),
        "month_options": _month_options(raw_months, year, month),
        "columns": first_columns,
        "first_rows": first_rows,
        "first_totals": first_totals,
        "first_grand": first_grand,
        "second_rows": second_rows,
        "second_totals": second_totals,
        "second_grand": second_grand,
        "export_url": export_url,
    }
    return render(request, "reviews/managing_partner/employee_summary.html", context)


_EXPENSE_BUCKET_LABELS = [
    ("mktg_meals", "Marketing - Meals"),
    ("mktg_ent", "Marketing - Entertainment"),
    ("mktg_gen", "Marketing - General"),
    ("recr_meals", "Recruiting - Meals"),
    ("recr_ent", "Recruiting - Entertainment"),
    ("recr_gen", "Recruiting - General"),
    ("ks_meals", "Keystone - Meals"),
    ("ks_ent", "Keystone - Entertainment"),
    ("ks_gen", "Keystone - General"),
    ("travel", "Travel"),
    ("office", "Office - Supplies"),
    ("computer", "Computer - Expenses"),
    ("training", "Training"),
    ("dues", "Dues & Subscriptions"),
    ("phone", "Telecom - Phone"),
    ("other", "Other"),
]


@login_required
@managing_partner_required
def employee_expenses(request):
    """
    Employee expenses summary: employees as rows, expense categories across
    the top with Client Expenses aggregated.
    """
    year, month = _parse_month_param(request.GET.get("month"))
    employees = _active_employees()

    columns = [("client_exp", "Client Expenses")]
    columns += list(_EXPENSE_BUCKET_LABELS)

    non_client_threshold = Decimal("1000")
    rows = []
    column_totals = {col_id: Decimal("0") for col_id, _ in columns}

    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        row_values = {}
        row_total = Decimal("0")
        expense_flags = {}

        if upload:
            expenses = upload.parsed_json.get("expenses", {})
            totals_by_bucket = expenses.get("totals_by_bucket", {})

            row_values["client_exp"] = Decimal(str(expenses.get("client_billed_total", 0)))

            for col_id, bucket_label in _EXPENSE_BUCKET_LABELS:
                amt = Decimal(str(totals_by_bucket.get(bucket_label, 0)))
                row_values[col_id] = amt
                if amt > non_client_threshold:
                    expense_flags[col_id] = True

        for col_id, _ in columns:
            v = row_values.get(col_id, Decimal("0"))
            column_totals[col_id] += v
            row_total += v

        rows.append({
            "employee": emp,
            "values": row_values,
            "total": row_total,
            "missing": upload is None,
            "expense_flags": expense_flags,
        })

    grand_total = sum(column_totals.values())

    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:12]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    context = {
        "month_label": _month_label(year, month),
        "month_options": _month_options(raw_months, year, month),
        "columns": columns,
        "rows": rows,
        "column_totals": column_totals,
        "grand_total": grand_total,
    }
    return render(request, "reviews/managing_partner/employee_expenses.html", context)


# =============================================================================
# PARTNER EXCEL EXPORT
# =============================================================================

@login_required
@managing_partner_required
def partner_export_xlsx(request, year, month):
    """Export the category and daily summary as a two-tab Excel workbook."""
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl not installed", status=500)

    wb = Workbook()

    employees = _active_employees()
    last_day = monthrange(year, month)[1]

    # --- Category Summary tab ---
    ws_cat = wb.active
    ws_cat.title = "Category Summary"
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    cat_headers = ["Category"] + [e.get_full_name() or e.email for e in employees] + ["Total"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    uploads_by_user = {}
    client_codes = set()
    client_labels = {}
    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        if upload:
            uploads_by_user[emp.id] = upload
            for half_key in ("first_half", "second_half"):
                for line in upload.parsed_json.get("time", {}).get(half_key, {}).get("lines", []):
                    if line.get("group") == "client" and line.get("charge_code"):
                        code = line["charge_code"]
                        client_codes.add(code)
                        if code not in client_labels:
                            client_labels[code] = line.get("label") or code

    client_order = list(ClientMapping.objects.filter(active=True).order_by("sort_order", "display_name"))
    client_order_codes = [c.code for c in client_order]
    ordered_client_codes = [c for c in client_order_codes if c in client_codes and c and c != "0"]
    ordered_client_codes += sorted([c for c in client_codes if c not in ordered_client_codes and c and c != "0"])

    marketing_rows_def = [
        ("Marketing - General/Other", "GEN"),
        ("Chicago - Strategics", "CHI-STRAT"),
        ("Chicago - Banks", "CHI-BNK"),
        ("Chicago - Existing", "CHI-EXST"),
        ("Chicago - PE", "CHI-PEG"),
        ("Atlanta - Strategics", "ATL-STRAT"),
        ("Atlanta - Banks", "ATL-BNK"),
        ("Atlanta - Existing", "ATL-EXST"),
        ("Atlanta - PE", "ATL-PEG"),
        ("Los Angeles - Strategics", "LAX-STRAT"),
        ("Los Angeles - Banks", "LAX-BNK"),
        ("Los Angeles - Existing", "LAX-EXST"),
        ("Los Angeles - PE", "LAX-PEG"),
    ]
    other_rows_def = [
        ("Administration", "ADM"),
        ("Recruiting", "REC"),
        ("Training", "TRN"),
        ("Meetings", "MTG"),
        ("PTO", "PTO"),
        ("Other Paid Time Off", "HOL+OFF"),
    ]

    def _write_cat_half(ws, start_row, half_key, half_label):
        row_idx = start_row
        section_font = Font(bold=True, size=10)
        ws.cell(row=row_idx, column=1, value=f"--- {half_label} ---").font = section_font
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="CLIENT WORK").font = section_font
        row_idx += 1

        for code in ordered_client_codes:
            mapping = next((c for c in client_order if c.code == code), None)
            label = mapping.display_name if mapping else client_labels.get(code, code)
            ws.cell(row=row_idx, column=1, value=label)
            row_total = Decimal("0")
            for ci, emp in enumerate(employees, 2):
                upload = uploads_by_user.get(emp.id)
                hours = Decimal("0")
                if upload:
                    hours = Decimal(str(upload.parsed_json.get("time", {}).get(half_key, {}).get("totals_by_client_code", {}).get(code, 0)))
                ws.cell(row=row_idx, column=ci, value=float(hours))
                row_total += hours
            ws.cell(row=row_idx, column=len(employees) + 2, value=float(row_total))
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="MARKETING").font = section_font
        row_idx += 1

        for label, code in marketing_rows_def:
            ws.cell(row=row_idx, column=1, value=label)
            row_total = Decimal("0")
            for ci, emp in enumerate(employees, 2):
                upload = uploads_by_user.get(emp.id)
                hours = Decimal("0")
                if upload:
                    hours = Decimal(str(upload.parsed_json.get("time", {}).get(half_key, {}).get("totals_by_marketing_bucket", {}).get(code, 0)))
                ws.cell(row=row_idx, column=ci, value=float(hours))
                row_total += hours
            ws.cell(row=row_idx, column=len(employees) + 2, value=float(row_total))
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="OTHER HOURS").font = section_font
        row_idx += 1

        for label, code in other_rows_def:
            ws.cell(row=row_idx, column=1, value=label)
            row_total = Decimal("0")
            for ci, emp in enumerate(employees, 2):
                upload = uploads_by_user.get(emp.id)
                hours = Decimal("0")
                if upload:
                    other_totals = upload.parsed_json.get("time", {}).get(half_key, {}).get("totals_by_other_hours", {})
                    if code == "HOL+OFF":
                        hours = Decimal(str(other_totals.get("HOL", 0))) + Decimal(str(other_totals.get("OFF", 0)))
                    else:
                        hours = Decimal(str(other_totals.get(code, 0)))
                ws.cell(row=row_idx, column=ci, value=float(hours))
                row_total += hours
            ws.cell(row=row_idx, column=len(employees) + 2, value=float(row_total))
            row_idx += 1

        return row_idx

    next_row = _write_cat_half(ws_cat, 2, "first_half", "Period One (Days 1–15)")
    next_row += 1
    _write_cat_half(ws_cat, next_row, "second_half", "Period Two (Days 16–End)")

    ws_cat.column_dimensions["A"].width = 30
    for ci in range(2, len(employees) + 3):
        ws_cat.column_dimensions[get_column_letter(ci)].width = 14

    # --- Daily Summary tab ---
    ws_daily = wb.create_sheet("Daily Summary")

    first_dates = [date(year, month, d) for d in range(1, 16)]
    second_dates = [date(year, month, d) for d in range(16, last_day + 1)]

    def _write_daily_half(ws, start_row, dates, half_key, half_label):
        row_idx = start_row
        section_font = Font(bold=True, size=10)
        ws.cell(row=row_idx, column=1, value=half_label).font = section_font
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="Employee").font = Font(bold=True)
        for di, d in enumerate(dates, 2):
            cell = ws.cell(row=row_idx, column=di, value=d.strftime("%m/%d"))
            cell.font = Font(bold=True)
        ws.cell(row=row_idx, column=len(dates) + 2, value="Total").font = Font(bold=True)
        row_idx += 1

        for emp in employees:
            ws.cell(row=row_idx, column=1, value=emp.get_full_name() or emp.email)
            upload = uploads_by_user.get(emp.id)
            total = Decimal("0")
            for di, d in enumerate(dates, 2):
                hours = Decimal("0")
                if upload:
                    hours = Decimal(str(upload.parsed_json.get("time", {}).get(half_key, {}).get("daily_totals", {}).get(d.isoformat(), 0)))
                ws.cell(row=row_idx, column=di, value=float(hours))
                total += hours
            ws.cell(row=row_idx, column=len(dates) + 2, value=float(total))
            row_idx += 1

        return row_idx

    next_row = _write_daily_half(ws_daily, 1, first_dates, "first_half", "Period One (Days 1–15)")
    next_row += 1
    _write_daily_half(ws_daily, next_row, second_dates, "second_half", "Period Two (Days 16–End)")

    ws_daily.column_dimensions["A"].width = 25
    for ci in range(2, max(len(first_dates), len(second_dates)) + 3):
        ws_daily.column_dimensions[get_column_letter(ci)].width = 10

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"partner_summary_{year}_{month:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# MANAGING PARTNER SPREADSHEET VIEW
# =============================================================================

MP_CANONICAL_PROJECTS = [
    "Bunn",
    "Remprex",
    "Ronnoco",
    "IPS - PMO Support",
    "IPS - Value Stream",
    "Weinstein",
    "Merit Brass",
    "Barrel Craft Spirits",
    "Breeze Autocare",
    "Human Active Technology",
    "Hayes Performance Systems",
    "AASC",
    "SNB",
    "Lion Equity",
    "H3 Mfg.",
    "Med Sol",
    "CHS",
    "MEC",
    "LeVecke - interim CFO",
    "Uncle Nearest",
    "Everde",
    "IPS - Eng. Org",
    "Sparkstone",
    "Royston - SCA Costing",
    "ASG AR recovery",
    "Chemline",
    "TPC",
    "Radwell",
    "Alliance Environmental",
    "Forte",
    "Golden State Foods",
    "Sauder",
    "Hansons",
    "Champion Homes",
    "Worthington",
    "Aquamar",
    "Bright Innovations",
    "MEC - Muskie",
    "Nexamotion",
    "Edelbrock",
    "ClarkDietrich",
    "Charcuterie Artisans",
    "Gerber",
    "IPS - Coil Strategy",
    "MEC - Mayflower",
    "Raydia",
    "AEM",
    "MEC - ERM",
    "Road Tested Parts",
    "Sauder - Plan B",
    "StyleCraft",
    "Hill & Smith",
    "Buddig",
    "SNB Cascade",
    "95PG",
    "Stifel - Tempest",
    "Pink Lily",
    "MGG - Sawmill",
    "WM Partners",
    "H3 - Interco Cost Savings",
    "GPS",
    "IPS - Edison",
]


def _mp_build_label_map(uploads_by_user):
    """Build a normalized label → charge_code mapping from all uploads."""
    label_to_code = {}
    for upload in uploads_by_user.values():
        for half_key in ("first_half", "second_half"):
            for line in upload.parsed_json.get("time", {}).get(half_key, {}).get("lines", []):
                if line.get("group") == "client" and line.get("charge_code"):
                    code = line["charge_code"]
                    label = (line.get("label") or "").strip()
                    if label:
                        label_to_code[label] = code
    return label_to_code


def _mp_project_active_map(ordered_projects, label_to_code, uploads_by_user):
    """Determine which projects have hours in either half (Active vs Inactive)."""
    active_map = {}
    for name in ordered_projects:
        code = label_to_code.get(name)
        has_hours = False
        if code:
            for upload in uploads_by_user.values():
                for hk in ("first_half", "second_half"):
                    val = upload.parsed_json.get("time", {}).get(hk, {}).get(
                        "totals_by_client_code", {}).get(code, 0)
                    if float(val) > 0:
                        has_hours = True
                        break
                if has_hours:
                    break
        active_map[name] = has_hours
    return active_map


def _build_mp_matrix(employees_ordered, uploads_by_user, ordered_projects,
                     label_to_code, project_active, half_key):
    """Build the matrix for one half of the Managing Partner spreadsheet."""
    marketing_rows_def = [
        ("Marketing - General/Other", "GEN"),
        ("Chicago - Strategics", "CHI-STRAT"),
        ("Chicago - Banks", "CHI-BNK"),
        ("Chicago - Existing", "CHI-EXST"),
        ("Chicago - PE", "CHI-PEG"),
        ("Atlanta - Strategics", "ATL-STRAT"),
        ("Atlanta - Banks", "ATL-BNK"),
        ("Atlanta - Existing", "ATL-EXST"),
        ("Atlanta - PE", "ATL-PEG"),
        ("Los Angeles - Strategics", "LAX-STRAT"),
        ("Los Angeles - Banks", "LAX-BNK"),
        ("Los Angeles - Existing", "LAX-EXST"),
        ("Los Angeles - PE", "LAX-PEG"),
    ]
    other_rows_def = [
        ("Administration", "ADM"),
        ("Recruiting", "REC"),
        ("Training", "TRN"),
        ("Meetings", "MTG"),
        ("PTO", "PTO"),
        ("Other Paid Time Off", "HOL+OFF"),
    ]

    matrix = []
    chargeable_totals = {emp.id: Decimal("0") for emp in employees_ordered}
    marketing_totals = {emp.id: Decimal("0") for emp in employees_ordered}
    other_totals = {emp.id: Decimal("0") for emp in employees_ordered}

    for project_name in ordered_projects:
        code = label_to_code.get(project_name)
        is_active = project_active.get(project_name, False)
        row = {"label": project_name, "code": code or "", "group": "client",
               "active": "Active" if is_active else "Inactive", "values": {}}
        total = Decimal("0")
        for emp in employees_ordered:
            upload = uploads_by_user.get(emp.id)
            hours = Decimal("0")
            if upload and code:
                hours = Decimal(str(upload.parsed_json.get("time", {}).get(
                    half_key, {}).get("totals_by_client_code", {}).get(code, 0)))
            row["values"][emp.id] = hours
            total += hours
            chargeable_totals[emp.id] += hours
        row["total"] = total
        matrix.append(row)

    matrix.append({
        "label": "Total Chargeable", "group": "total_chargeable",
        "values": dict(chargeable_totals), "total": sum(chargeable_totals.values()),
    })

    for label, code in marketing_rows_def:
        row = {"label": label, "code": code, "group": "marketing", "values": {}}
        total = Decimal("0")
        for emp in employees_ordered:
            upload = uploads_by_user.get(emp.id)
            hours = Decimal("0")
            if upload:
                hours = Decimal(str(upload.parsed_json.get("time", {}).get(
                    half_key, {}).get("totals_by_marketing_bucket", {}).get(code, 0)))
            row["values"][emp.id] = hours
            total += hours
            marketing_totals[emp.id] += hours
        row["total"] = total
        matrix.append(row)

    matrix.append({
        "label": "Total Marketing", "group": "total_marketing",
        "values": dict(marketing_totals), "total": sum(marketing_totals.values()),
    })

    for label, code in other_rows_def:
        row = {"label": label, "code": code, "group": "other", "values": {}}
        total = Decimal("0")
        for emp in employees_ordered:
            upload = uploads_by_user.get(emp.id)
            hours = Decimal("0")
            if upload:
                other_totals_data = upload.parsed_json.get("time", {}).get(
                    half_key, {}).get("totals_by_other_hours", {})
                if code == "HOL+OFF":
                    hours = (Decimal(str(other_totals_data.get("HOL", 0)))
                             + Decimal(str(other_totals_data.get("OFF", 0))))
                else:
                    hours = Decimal(str(other_totals_data.get(code, 0)))
            row["values"][emp.id] = hours
            total += hours
            other_totals[emp.id] += hours
        row["total"] = total
        matrix.append(row)

    matrix.append({
        "label": "Total Other Hours", "group": "total_other",
        "values": dict(other_totals), "total": sum(other_totals.values()),
    })

    grand = {eid: chargeable_totals[eid] + marketing_totals[eid] + other_totals[eid]
             for eid in chargeable_totals}
    matrix.append({
        "label": "Total Hours", "group": "grand_total",
        "values": grand, "total": sum(grand.values()),
    })

    return matrix


def _mp_common_data(year, month):
    """Shared data-loading logic for the MP spreadsheet view and export."""
    employees = _active_employees()
    uploads_by_user = {}
    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        if upload:
            uploads_by_user[emp.id] = upload

    label_to_code = _mp_build_label_map(uploads_by_user)

    layout, _ = ManagingPartnerLayout.objects.get_or_create(
        year=year, month=month,
        defaults={"employee_order": [], "client_order": []}
    )

    emp_map = {emp.id: emp for emp in employees}
    active_ids = set(emp_map.keys())

    if layout.employee_order:
        ordered_ids = [eid for eid in layout.employee_order if eid in active_ids]
        for emp in employees:
            if emp.id not in ordered_ids:
                ordered_ids.append(emp.id)
        employees_ordered = [emp_map[eid] for eid in ordered_ids if eid in emp_map]
    else:
        employees_ordered = list(employees)
        layout.employee_order = [emp.id for emp in employees_ordered]

    if layout.client_order:
        seen = set(layout.client_order)
        ordered_projects = list(layout.client_order)
        for name in MP_CANONICAL_PROJECTS:
            if name not in seen:
                ordered_projects.append(name)
                seen.add(name)
    else:
        ordered_projects = list(MP_CANONICAL_PROJECTS)
        layout.client_order = ordered_projects

    layout.save(update_fields=["employee_order", "client_order", "updated_at"])

    project_active = _mp_project_active_map(ordered_projects, label_to_code, uploads_by_user)

    first_matrix = _build_mp_matrix(
        employees_ordered, uploads_by_user, ordered_projects,
        label_to_code, project_active, "first_half",
    )
    second_matrix = _build_mp_matrix(
        employees_ordered, uploads_by_user, ordered_projects,
        label_to_code, project_active, "second_half",
    )
    return employees_ordered, first_matrix, second_matrix


@login_required
@managing_partner_only_required
def mp_spreadsheet(request):
    """Managing Partner spreadsheet view replicating the Excel layout."""
    year, month = _parse_month_param(request.GET.get("month"))
    employees_ordered, first_matrix, second_matrix = _mp_common_data(year, month)

    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:12]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    from django.urls import reverse
    export_url = reverse("reviews:mp_export_xlsx", args=[year, month])

    context = {
        "employees": employees_ordered,
        "month_label": _month_label(year, month),
        "month_param": f"{year}-{month:02d}",
        "month_options": _month_options(raw_months, year, month),
        "first_matrix": first_matrix,
        "second_matrix": second_matrix,
        "export_url": export_url,
    }
    return render(request, "reviews/managing_partner/mp_spreadsheet.html", context)


@login_required
@managing_partner_only_required
@require_POST
def mp_reorder(request):
    """AJAX endpoint to persist drag-and-drop reorder changes."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    year = data.get("year")
    month = data.get("month")
    field = data.get("field")
    order = data.get("order")

    if not all([year, month, field, order]) or field not in ("employee_order", "client_order"):
        return JsonResponse({"error": "Bad request"}, status=400)

    layout, _ = ManagingPartnerLayout.objects.get_or_create(
        year=int(year), month=int(month),
        defaults={"employee_order": [], "client_order": []}
    )

    if field == "employee_order":
        layout.employee_order = [int(x) for x in order]
    else:
        layout.client_order = list(order)

    layout.updated_by = request.user
    layout.save()
    return JsonResponse({"ok": True})


@login_required
@managing_partner_only_required
def mp_export_xlsx(request, year, month):
    """Export the Managing Partner spreadsheet as Excel, respecting custom ordering."""
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl not installed", status=500)

    employees_ordered, first_matrix, second_matrix = _mp_common_data(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Time Collection"
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_font = Font(bold=True, size=10)
    section_font = Font(bold=True, size=10, color="333333")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    ws.cell(row=1, column=1, value="Time Collection").font = Font(bold=True, size=14)

    emp_initials = []
    for emp in employees_ordered:
        initials = emp.profile_or_none.initials if emp.profile_or_none else emp.first_name[:2].upper()
        emp_initials.append(initials or emp.first_name[:2].upper())

    ws.cell(row=3, column=2, value="Active / Inactive").font = header_font
    ws.cell(row=3, column=2).fill = header_fill
    for ci, init in enumerate(emp_initials, 3):
        cell = ws.cell(row=3, column=ci, value=init)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=len(emp_initials) + 3, value="Total").font = header_font
    ws.cell(row=3, column=len(emp_initials) + 3).fill = header_fill

    def write_half(start_row, matrix, half_label):
        row_idx = start_row
        ws.cell(row=row_idx, column=1, value=half_label).font = Font(bold=True, size=12)
        row_idx += 2

        for entry in matrix:
            label = entry["label"]
            group = entry.get("group", "")

            if group == "total_chargeable":
                row_idx += 1

            if group == "marketing" and entry.get("code") == "GEN":
                row_idx += 1
                ws.cell(row=row_idx, column=1, value="Marketing").font = section_font
                ws.cell(row=row_idx, column=1).fill = section_fill
                row_idx += 1

            if group == "other" and entry.get("code") == "ADM":
                row_idx += 1
                ws.cell(row=row_idx, column=1, value="Other Hours").font = section_font
                ws.cell(row=row_idx, column=1).fill = section_fill
                row_idx += 1

            is_total = group in ("total_chargeable", "total_marketing", "total_other", "grand_total")
            ws.cell(row=row_idx, column=1, value=label)
            if is_total:
                ws.cell(row=row_idx, column=1).font = total_font

            if group == "client":
                ws.cell(row=row_idx, column=2, value=entry.get("active", "Active"))

            for ci, emp in enumerate(employees_ordered, 3):
                val = float(entry["values"].get(emp.id, 0))
                cell = ws.cell(row=row_idx, column=ci, value=val if val else None)
                cell.alignment = Alignment(horizontal="center")
                if is_total:
                    cell.font = total_font

            total_val = float(entry.get("total", 0))
            total_cell = ws.cell(
                row=row_idx, column=len(employees_ordered) + 3,
                value=total_val if total_val else None,
            )
            if is_total:
                total_cell.font = total_font

            row_idx += 1

        return row_idx

    next_row = write_half(4, first_matrix, "Period One")
    next_row += 2
    write_half(next_row, second_matrix, "Period Two")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    for ci in range(3, len(employees_ordered) + 4):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"managing_partner_{year}_{month:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# PAYROLL PARTNER VIEWS
# =============================================================================

@login_required
@payroll_partner_required
def payroll_dashboard(request):
    """Partner dashboard."""
    year, month = _parse_month_param(request.GET.get("month"))
    raw_months = (
        TimesheetUpload.objects.values_list("year", "month")
        .distinct()
        .order_by("-year", "-month")[:6]
    )
    if (year, month) not in raw_months:
        raw_months = [(year, month)] + list(raw_months)

    employees = _active_employees()
    expense_summary = []
    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        if upload:
            total = upload.parsed_json.get("expenses", {}).get("total_expenses", 0)
            status = upload.status
        else:
            total = 0
            status = "MISSING"
        expense_summary.append({
            "employee": emp,
            "upload": upload,
            "total": total,
            "status": status,
        })

    total_expenses = sum(
        Decimal(str(item.get("total", 0))) for item in expense_summary
    )
    report_count = sum(1 for item in expense_summary if item.get("upload"))

    context = {
        "month_options": _month_options(raw_months, year, month),
        "selected_month": f"{year}-{month:02d}",
        "selected_month_label": _month_label(year, month),
        "expense_summary": expense_summary,
        "total_expenses": total_expenses,
        "report_count": report_count,
    }
    return render(request, "reviews/payroll/dashboard.html", context)


@login_required
@payroll_partner_required
def payroll_export(request, year, month):
    format_type = request.GET.get("format", "xlsx")
    rows, flags = _build_payroll_rows(year, month)

    if format_type == "flags":
        return _render_flags_csv(flags, year, month)
    if format_type == "csv":
        return _render_payroll_csv(rows, year, month)
    return _render_payroll_xlsx(rows, year, month)


PAYROLL_COLUMNS = [
    "Person",
    "Initials",
    "EE#",
    "Reimbursed",
    "MKT Banking — Chicago",
    "MKT Banking — Chicago-Lead",
    "MKT Banking — Atlanta",
    "MKT Banking — Atlanta-Lead",
    "MKT Banking — LAX",
    "MKT Banking — LAX-Lead",
    "MKT Existing — Chicago",
    "MKT Existing — Chicago-Lead",
    "MKT Existing — Atlanta",
    "MKT Existing — Atlanta-Lead",
    "MKT Existing — LAX",
    "MKT Existing — LAX-Lead",
    "MKT General — General",
    "MKT General — General-Lead",
    "MKT Strategics — Chicago",
    "MKT Strategics — Chicago-Lead",
    "MKT Strategics — Atlanta",
    "MKT Strategics — Atlanta-Lead",
    "MKT Strategics — LAX",
    "MKT Strategics — LAX-Lead",
    "MKT PEG — Chicago",
    "MKT PEG — Chicago-Lead",
    "MKT PEG — Atlanta",
    "MKT PEG — Atlanta-Lead",
    "MKT PEG — LAX",
    "MKT PEG — LAX-Lead",
    "Recr. — Meals",
    "Recr. — Entertainment",
    "Recr. — General",
    "Training — Expenses",
    "TKG — Meals",
    "TKG — Entertainment",
    "TKG — General",
    "TKG — Travel",
    "Office — Supplies",
    "Computer — Expenses",
    "Dues & Subscriptions",
    "Phone",
    "Other",
    "Expenses — Total",
    "Front Page — Reimb.",
]


NON_MARKETING_BUCKET_MAP = {
    "Recruiting - Meals": "Recr. — Meals",
    "Recruiting - Entertainment": "Recr. — Entertainment",
    "Recruiting - General": "Recr. — General",
    "Training": "Training — Expenses",
    "Keystone - Meals": "TKG — Meals",
    "Keystone - Entertainment": "TKG — Entertainment",
    "Keystone - General": "TKG — General",
    "Travel": "TKG — Travel",
    "Office - Supplies": "Office — Supplies",
    "Computer - Expenses": "Computer — Expenses",
    "Dues & Subscriptions": "Dues & Subscriptions",
    "Telecom - Phone": "Phone",
    "Other": "Other",
}


def _build_payroll_rows(year, month):
    employees = _active_employees()
    rows = []
    flags = []
    threshold = Decimal(str(getattr(settings, "PAYROLL_FLAG_CELL_THRESHOLD", 500.0)))

    for emp in employees:
        upload = _latest_upload_for_user(emp, year, month)
        row = {col: Decimal("0") for col in PAYROLL_COLUMNS}
        row["Person"] = emp.get_full_name() or emp.email
        row["Initials"] = emp.profile_or_none.initials if emp.profile_or_none else ""
        row["EE#"] = emp.profile_or_none.employee_number if emp.profile_or_none else ""

        employee_flags = []
        if not upload:
            employee_flags.append("MISSING_SUBMISSION")
            rows.append(row)
            flags.append(_flag_row(emp, year, month, employee_flags))
            continue

        if upload.has_blocking_errors:
            employee_flags.append("HAS_BLOCKING_ERRORS")

        expenses = upload.parsed_json.get("expenses", {})
        totals_by_bucket = expenses.get("totals_by_bucket", {})
        items = expenses.get("items", [])

        # Non-marketing buckets
        for bucket, target in NON_MARKETING_BUCKET_MAP.items():
            row[target] = Decimal(str(totals_by_bucket.get(bucket, 0)))

        # Marketing allocations from items
        unclassified = Decimal("0")
        for item in items:
            bucket = item.get("bucket")
            if not bucket or not bucket.startswith("Marketing"):
                continue
            amount = Decimal(str(item.get("amount", 0)))
            code = item.get("charge_code") or ""
            column = _marketing_column_for_code(code)
            if column:
                row[column] += amount
            else:
                unclassified += amount

        if unclassified > 0:
            employee_flags.append("UNCLASSIFIED_MARKETING_CODE")

        # Totals
        expense_total = sum(row[col] for col in PAYROLL_COLUMNS if col not in ["Person", "Initials", "EE#", "Reimbursed", "Expenses — Total", "Front Page — Reimb."])
        row["Expenses — Total"] = expense_total
        row["Reimbursed"] = expense_total
        row["Front Page — Reimb."] = expense_total

        for col in PAYROLL_COLUMNS:
            if isinstance(row[col], Decimal) and row[col] > threshold:
                employee_flags.append(f"CELL_ABOVE_THRESHOLD:{col}")

        rows.append(row)
        if employee_flags:
            flags.append(_flag_row(emp, year, month, employee_flags))

    return rows, flags


def _marketing_column_for_code(code):
    code = (code or "").upper()
    if not code:
        return None

    lead = "Lead" if "-LEAD" in code else "Other"
    region = None
    segment = None

    if code.startswith("GEN"):
        segment = "General"
    elif "-BNK" in code:
        segment = "Banking"
    elif "-EXST" in code:
        segment = "Existing"
    elif "-STRAT" in code:
        segment = "Strategics"
    elif "-PEG" in code:
        segment = "PEG"

    if code.startswith("CHI"):
        region = "Chicago"
    elif code.startswith("ATL"):
        region = "Atlanta"
    elif code.startswith("LAX"):
        region = "LAX"

    if not segment:
        return None
    if segment == "General":
        return f"MKT General — General{'' if lead == 'Other' else '-Lead'}"
    if not region:
        return None

    label = f"MKT {segment} — {region}"
    if lead == "Lead":
        label = f"{label}-Lead"
    return label


def _render_payroll_csv(rows, year, month):
    response = HttpResponse(content_type="text/csv")
    filename = f"payroll_expenses_{year}_{month:02d}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(PAYROLL_COLUMNS)
    for row in rows:
        writer.writerow([_csv_value(row.get(col)) for col in PAYROLL_COLUMNS])
    return response


def _render_payroll_xlsx(rows, year, month):
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl not installed", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    header_font = Font(bold=True, size=11)
    for col_idx, col in enumerate(PAYROLL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(PAYROLL_COLUMNS, start=1):
            value = row.get(col)
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"payroll_expenses_{year}_{month:02d}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _render_flags_csv(flags, year, month):
    response = HttpResponse(content_type="text/csv")
    filename = f"payroll_flags_{year}_{month:02d}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Email", "Name", "Month", "Flags"])
    for row in flags:
        writer.writerow([row["email"], row["name"], row["month"], "; ".join(row["flags"])])
    return response


def _flag_row(emp, year, month, flags):
    return {
        "email": emp.email,
        "name": emp.get_full_name() or emp.email,
        "month": f"{year}-{month:02d}",
        "flags": flags,
    }


def _csv_value(value):
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    return value
