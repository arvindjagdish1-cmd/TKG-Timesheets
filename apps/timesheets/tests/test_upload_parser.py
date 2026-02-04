from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook

from apps.timesheets.services.upload_parser import parse_timesheet_workbook
from apps.timesheets.services.upload_validation import validate_parsed_workbook


class UploadParserTests(TestCase):
    def _build_workbook(self):
        wb = Workbook()
        wb.remove(wb.active)

        wb.create_sheet("Time-1st half of month")
        wb.create_sheet("Time-2nd half of month")
        wb.create_sheet("Expenses-Main")
        wb.create_sheet("Expenses-Additional")
        wb.create_sheet("Auto Log 655")
        wb.create_sheet("Validations")

        time1 = wb["Time-1st half of month"]
        time2 = wb["Time-2nd half of month"]
        validations = wb["Validations"]
        expenses = wb["Expenses-Main"]

        time1["A1"] = "The Keystone Group"
        time1["L1"] = "Test User"
        time1["T1"] = 2026
        time1["V1"] = 1
        time1["A39"] = "Version 01.19.2024"

        validations["A1"] = "General"
        validations["B1"] = "GEN"

        time1["A6"] = "Client A"
        time1["U6"] = "CLIENT1"
        time1["B6"] = 8

        time2["A6"] = "Client A"
        time2["U6"] = "CLIENT1"
        time2["B6"] = 7.5

        time1["A16"] = "General"
        time1["B16"] = 1

        expenses["A6"] = "2026-01-05"
        expenses["B6"] = "Lunch"
        expenses["V6"] = "GEN-LEAD"
        expenses["E6"] = 10

        expenses["A7"] = "2026-01-06"
        expenses["B7"] = "Supplies"
        expenses["V7"] = "ADM"
        expenses["O7"] = 25

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_parse_and_validate(self):
        file_bytes = self._build_workbook()
        parsed = parse_timesheet_workbook(file_bytes)

        self.assertEqual(parsed["period"]["year"], 2026)
        self.assertEqual(parsed["period"]["month"], 1)
        self.assertGreater(parsed["time"]["first_half"]["total_hours"], 0)
        self.assertGreater(parsed["expenses"]["total_expenses"], 0)

        issues = validate_parsed_workbook(parsed)
        blocking = [i for i in issues if i["severity"] == "ERROR"]
        self.assertEqual(blocking, [])

    def test_missing_charge_code_error(self):
        file_bytes = self._build_workbook()
        parsed = parse_timesheet_workbook(file_bytes)
        parsed["time"]["first_half"]["lines"][0]["charge_code"] = ""

        issues = validate_parsed_workbook(parsed)
        codes = {issue["code"] for issue in issues}
        self.assertIn("TIME_MISSING_CHARGE_CODE", codes)
