import calendar
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


REQUIRED_SHEETS = [
    "Time-1st half of month",
    "Time-2nd half of month",
    "Expenses-Main",
    "Expenses-Additional",
    "Auto Log 655",
    "Validations",
]

EXPENSE_BUCKET_COLUMNS = [
    ("E", "Marketing - Meals"),
    ("F", "Marketing - Entertainment"),
    ("G", "Marketing - General"),
    ("H", "Recruiting - Meals"),
    ("I", "Recruiting - Entertainment"),
    ("J", "Recruiting - General"),
    ("K", "Keystone - Meals"),
    ("L", "Keystone - Entertainment"),
    ("M", "Keystone - General"),
    ("N", "Travel"),
    ("O", "Office - Supplies"),
    ("P", "Computer - Expenses"),
    ("Q", "Training"),
    ("R", "Dues & Subscriptions"),
    ("S", "Telecom - Phone"),
    ("T", "Other"),
]


def parse_timesheet_workbook(file_bytes):
    wb_data = load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )
    wb_formulas = load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=False,
    )

    sheets_present = wb_data.sheetnames
    ws_data = {name: wb_data[name] for name in wb_data.sheetnames}
    ws_formulas = {name: wb_formulas[name] for name in wb_formulas.sheetnames}

    time_first_data = ws_data.get("Time-1st half of month")
    time_first_formulas = ws_formulas.get("Time-1st half of month")
    time_second_data = ws_data.get("Time-2nd half of month")
    time_second_formulas = ws_formulas.get("Time-2nd half of month")
    validations_ws = ws_data.get("Validations")

    metadata = _parse_metadata(time_first_data)
    year = metadata.get("year")
    month = metadata.get("month")

    validations_map = _parse_validations_map(validations_ws)

    time_data = {
        "first_half": _parse_time_half(
            time_first_data, "FIRST", year, month, validations_map,
            ws_formulas=time_first_formulas,
            all_ws_data=ws_data,
        ),
        "second_half": _parse_time_half(
            time_second_data, "SECOND", year, month, validations_map,
            ws_formulas=time_second_formulas,
            all_ws_data=ws_data,
        ),
    }

    expenses_data = _parse_expenses(ws_data)
    mileage_data = _parse_mileage(ws_data.get("Auto Log 655"))

    return {
        "sheets_present": sheets_present,
        "metadata": metadata,
        "period": {"year": year, "month": month},
        "time": time_data,
        "expenses": expenses_data,
        "mileage": mileage_data,
    }


def _parse_metadata(ws):
    if not ws:
        return {}
    return {
        "company": _string_cell(ws, "A1"),
        "employee_name": _string_cell(ws, "L1"),
        "year": _int_cell(ws, "T1"),
        "month": _int_cell(ws, "V1"),
        "mid_marker": _int_cell(ws, "X1"),
        "template_version": _string_cell(ws, "A39"),
    }


def _parse_validations_map(ws):
    if not ws:
        return {}
    mapping = {}
    for row in ws.iter_rows(min_row=1, max_col=2):
        category = _to_string(row[0].value)
        base_code = _to_string(row[1].value)
        if category and base_code:
            mapping[category] = base_code
    return mapping


def _parse_time_half(ws, half, year, month, validations_map,
                     ws_formulas=None, all_ws_data=None):
    if not ws or not year or not month:
        return _empty_time_half()

    day_range = _get_day_range(year, month, half)
    date_columns = _map_day_columns(half, day_range)

    lines = []
    totals_by_client_code = {}
    totals_by_marketing_bucket = {}
    totals_by_other_hours = {}
    daily_totals = {d.isoformat(): Decimal("0") for d in day_range}

    # Client rows
    for row_idx in range(6, 14):
        line = _parse_time_row(
            ws,
            row_idx,
            date_columns,
            group="client",
            label_cell="A",
            code_cell="U",
            ws_formulas=ws_formulas,
            all_ws_data=all_ws_data,
        )
        lines.append(line)
        _accumulate_time_totals(
            line,
            daily_totals,
            totals_by_client_code,
            code_key=line.get("charge_code"),
        )

    # Marketing rows
    for row_idx in range(16, 30):
        category = _resolve_cell(ws, ws_formulas, all_ws_data, f"A{row_idx}", as_string=True)
        base_code = validations_map.get(category) if category else ""
        line = _parse_time_row(
            ws,
            row_idx,
            date_columns,
            group="marketing",
            label=category,
            charge_code=base_code,
            category=category,
            ws_formulas=ws_formulas,
            all_ws_data=all_ws_data,
        )
        lines.append(line)
        _accumulate_time_totals(
            line,
            daily_totals,
            totals_by_marketing_bucket,
            code_key=base_code,
        )

    # Internal rows
    for row_idx in range(30, 37):
        line = _parse_time_row(
            ws,
            row_idx,
            date_columns,
            group="internal",
            label_cell="A",
            code_cell="U",
            ws_formulas=ws_formulas,
            all_ws_data=all_ws_data,
        )
        lines.append(line)
        _accumulate_time_totals(
            line,
            daily_totals,
            totals_by_other_hours,
            code_key=line.get("charge_code"),
        )

    total_hours = sum(daily_totals.values())

    return {
        "dates": [d.isoformat() for d in day_range],
        "daily_totals": _serialize_decimal_map(daily_totals),
        "lines": lines,
        "totals_by_client_code": _serialize_decimal_map(totals_by_client_code),
        "totals_by_marketing_bucket": _serialize_decimal_map(totals_by_marketing_bucket),
        "totals_by_other_hours": _serialize_decimal_map(totals_by_other_hours),
        "total_hours": float(total_hours),
    }


def _parse_time_row(
    ws,
    row_idx,
    date_columns,
    group,
    label_cell=None,
    code_cell=None,
    label=None,
    charge_code=None,
    category=None,
    ws_formulas=None,
    all_ws_data=None,
):
    if label_cell:
        label = _resolve_cell(ws, ws_formulas, all_ws_data,
                              f"{label_cell}{row_idx}", as_string=True)
    if code_cell:
        charge_code = _resolve_cell(ws, ws_formulas, all_ws_data,
                                    f"{code_cell}{row_idx}", as_string=True)

    hours_by_day = {}
    row_total = Decimal("0")

    for day, column in date_columns.items():
        value = _decimal_cell(ws, f"{column}{row_idx}")
        hours_by_day[day.isoformat()] = float(value)
        row_total += value

    return {
        "row": row_idx,
        "group": group,
        "label": label,
        "category": category,
        "charge_code": charge_code,
        "hours_by_day": hours_by_day,
        "row_total": float(row_total),
    }


def _accumulate_time_totals(line, daily_totals, totals_by_code, code_key):
    for day_str, hours in line.get("hours_by_day", {}).items():
        daily_totals[day_str] = daily_totals.get(day_str, Decimal("0")) + Decimal(str(hours))
    if code_key:
        totals_by_code[code_key] = totals_by_code.get(code_key, Decimal("0")) + Decimal(
            str(line.get("row_total", 0))
        )


def _parse_expenses(worksheets):
    totals_by_bucket = {label: Decimal("0") for _, label in EXPENSE_BUCKET_COLUMNS}
    totals_by_charge_code = {}
    items = []

    marketing_total = Decimal("0")
    keystone_paid_total = Decimal("0")
    client_billed_total = Decimal("0")

    for sheet_name in ["Expenses-Main", "Expenses-Additional"]:
        ws = worksheets.get(sheet_name)
        if not ws:
            continue
        for row_idx in range(5, 39):
            date_val = _parse_date(ws.cell(row=row_idx, column=1).value)
            description = _to_string(ws.cell(row=row_idx, column=2).value)
            misc_value = ws.cell(row=row_idx, column=3).value
            d_value = ws.cell(row=row_idx, column=4).value
            charge_code = _string_cell(ws, f"V{row_idx}")

            bucket_amounts = []
            row_amount_total = Decimal("0")
            row_marketing_total = Decimal("0")

            for col, label in EXPENSE_BUCKET_COLUMNS:
                amount = _decimal_cell(ws, f"{col}{row_idx}")
                if amount > 0:
                    bucket_amounts.append((label, amount))
                    totals_by_bucket[label] += amount
                    row_amount_total += amount
                    keystone_paid_total += amount
                    if label.startswith("Marketing"):
                        row_marketing_total += amount

            numeric_d = _decimal_value(d_value)
            if numeric_d is not None:
                client_billed_total += numeric_d

            if _row_is_active(date_val, description, row_amount_total, charge_code):
                for label, amount in bucket_amounts:
                    items.append({
                        "sheet": sheet_name,
                        "row": row_idx,
                        "date": date_val.isoformat() if date_val else None,
                        "description": description,
                        "charge_code": charge_code,
                        "bucket": label,
                        "amount": float(amount),
                        "misc": _to_string(misc_value),
                        "client_billed": float(numeric_d) if numeric_d is not None else None,
                    })
                    if charge_code:
                        totals_by_charge_code[charge_code] = totals_by_charge_code.get(
                            charge_code, Decimal("0")
                        ) + amount

            marketing_total += row_marketing_total

    total_expenses = keystone_paid_total + client_billed_total

    return {
        "items": items,
        "totals_by_bucket": _serialize_decimal_map(totals_by_bucket),
        "totals_by_charge_code": _serialize_decimal_map(totals_by_charge_code),
        "marketing_total": float(marketing_total),
        "keystone_paid_total": float(keystone_paid_total),
        "client_billed_total": float(client_billed_total),
        "total_expenses": float(total_expenses),
    }


def _parse_mileage(ws):
    if not ws:
        return {"entries": [], "totals": {"miles_driven": 0, "net_miles": 0}}

    entries = []
    total_miles = Decimal("0")
    total_net = Decimal("0")

    for row_idx in range(7, ws.max_row + 1):
        date_val = _parse_date(ws.cell(row=row_idx, column=1).value)
        destination = _to_string(ws.cell(row=row_idx, column=3).value)
        odometer_start = _decimal_value(ws.cell(row=row_idx, column=5).value)
        odometer_end = _decimal_value(ws.cell(row=row_idx, column=6).value)
        commute = _decimal_value(ws.cell(row=row_idx, column=9).value) or Decimal("0")

        if not any([date_val, destination, odometer_start, odometer_end, commute]):
            continue

        miles_driven = Decimal("0")
        if odometer_start is not None and odometer_end is not None:
            miles_driven = max(Decimal("0"), odometer_end - odometer_start)

        net_miles = miles_driven - commute if commute else miles_driven

        entries.append({
            "row": row_idx,
            "date": date_val.isoformat() if date_val else None,
            "destination": destination,
            "odometer_start": float(odometer_start) if odometer_start is not None else None,
            "odometer_end": float(odometer_end) if odometer_end is not None else None,
            "commute_miles": float(commute) if commute else 0,
            "miles_driven": float(miles_driven),
            "net_miles": float(net_miles),
        })

        total_miles += miles_driven
        total_net += net_miles

    return {
        "entries": entries,
        "totals": {"miles_driven": float(total_miles), "net_miles": float(total_net)},
    }


def _empty_time_half():
    return {
        "dates": [],
        "daily_totals": {},
        "lines": [],
        "totals_by_client_code": {},
        "totals_by_marketing_bucket": {},
        "totals_by_other_hours": {},
        "total_hours": 0,
    }


def _get_day_range(year, month, half):
    last_day = calendar.monthrange(year, month)[1]
    if half == "FIRST":
        start_day, end_day = 1, 15
    else:
        start_day, end_day = 16, last_day
    return [date(year, month, day) for day in range(start_day, end_day + 1)]


def _map_day_columns(half, day_range):
    columns = {}
    if not day_range:
        return columns
    if half == "FIRST":
        base_ord = ord("B")
        for idx, d in enumerate(day_range[:15]):
            columns[d] = chr(base_ord + idx)
    else:
        base_ord = ord("B")
        for idx, d in enumerate(day_range):
            columns[d] = chr(base_ord + idx)
    return columns


def _row_is_active(date_val, description, row_amount_total, charge_code):
    return bool(date_val or description or row_amount_total > 0 or charge_code)


import re as _re

_CROSS_SHEET_RE = _re.compile(
    r"^='?([^'!]+)'?!([A-Z]+)(\d+)$"
)


def _resolve_cell(ws_data, ws_formulas, all_ws_data, cell_ref, as_string=False):
    """
    Return the resolved value of a cell.

    With data_only=True, openpyxl returns the cached computed value.
    If that is None/empty and the formula worksheet shows a cross-sheet
    reference like ='Time-1st half of month'!U6, we follow the reference
    and read the value from the source sheet (also opened data_only).
    """
    data_val = ws_data[cell_ref].value if ws_data else None

    if data_val is not None and str(data_val).strip() != "":
        return _to_string(data_val) if as_string else data_val

    if ws_formulas is not None:
        formula_val = ws_formulas[cell_ref].value
        if isinstance(formula_val, str) and formula_val.startswith("="):
            m = _CROSS_SHEET_RE.match(formula_val)
            if m and all_ws_data:
                ref_sheet, ref_col, ref_row = m.group(1), m.group(2), m.group(3)
                source_ws = all_ws_data.get(ref_sheet)
                if source_ws:
                    resolved = source_ws[f"{ref_col}{ref_row}"].value
                    if resolved is not None:
                        return _to_string(resolved) if as_string else resolved

            if as_string:
                return ""

    return _to_string(data_val) if as_string else data_val


def _decimal_cell(ws, cell_ref):
    val = ws[cell_ref].value
    if isinstance(val, str) and (val.startswith("=") or val.startswith("'")):
        return Decimal("0")
    return _decimal_value(val) or Decimal("0")


def _string_cell(ws, cell_ref):
    val = ws[cell_ref].value
    s = _to_string(val)
    if s.startswith("=") or s.startswith("'"):
        return ""
    return s


def _int_cell(ws, cell_ref):
    value = ws[cell_ref].value
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _to_string(value):
    if value is None:
        return ""
    return str(value).strip()


def _decimal_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _serialize_decimal_map(data):
    return {key: float(value) for key, value in data.items()}


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    return None
